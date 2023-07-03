""" Here we store Excel utilities """
import openpyxl

from calorie_count.src.DB.food_db import FoodDB, Food
from calorie_count.src.DB.meal_entry_db import MealEntryDB, MealEntry

DEFAULT_XLSX = 'Calorie_Counting.xlsx'
FOOD_SHEET = 'My Foods'
MEALS_SHEET = 'My Meal Entries'


def save_to_excel(path: str = DEFAULT_XLSX, *args) -> None:
    """Save Foods and entries to xlsx file
    The file will have 2 sheets: 'My Foods' and 'My Meal Entries'"""
    wb = openpyxl.Workbook()

    # --1-- Creating Foods sheet
    wb.active.title = FOOD_SHEET
    with FoodDB() as fdb:
        foods = fdb.get_all_foods()
    if foods:
        wb.active.append(foods[0].columns())
        for food in foods:
            wb.active.append(food.values)

    # --2-- Creating meals sheet
    sh = wb.create_sheet(MEALS_SHEET)
    with MealEntryDB() as mdb:
        start, end = map(str, mdb.get_first_last_dates())
        entries = mdb.get_entries_between_dates(start, end)
        print(entries)
    if entries:
        sh.append(MealEntry.columns())
        for entry in entries:
            sh.append(entry.values)

    # --3-- Saving Workbook
    print(f'Saving file here:', path)
    wb.save(path)


def import_excel(path: str = DEFAULT_XLSX, *args) -> None:
    """Load Foods and entries from xlsx file
    The file must have 2 sheets: 'My Foods' and 'My Meal Entries'
    The entries are added to the existing entries."""
    wb = openpyxl.load_workbook(path)

    # --1-- Reading Foods sheet

    gen = wb[FOOD_SHEET].iter_rows(values_only=True)
    headers = next(gen)
    assert headers == Food.columns(), f'Invalid Sheet: {FOOD_SHEET}\n' \
                                      f'Expected: {Food.columns}\nGot: {headers}'
    with FoodDB() as fdb:
        for row in gen:
            food = Food(*row)
            fdb.add_food(food, update=True)

    # --2-- Reading meals sheet
    gen = wb[MEALS_SHEET].iter_rows(values_only=True)
    headers = next(gen)
    assert headers == MealEntry.columns(), f'Invalid Sheet: {MEALS_SHEET}\n' \
                                           f'Expected: {MealEntry.columns}\nGot: {headers}'
    with MealEntryDB() as mdb:
        for row in gen:
            date, name, portion, *_ = row
            entry = MealEntry(name=name, date=date, portion=portion)
            mdb.add_meal_entry(entry)
    print(f'{path} Loaded!')


if __name__ == '__main__':
    save_to_excel()
    import_excel()
